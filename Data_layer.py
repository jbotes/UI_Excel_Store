
import openpyxl
import pprint as pp

# https://www.youtube.com/watch?v=ibXT3SbfkOc&list=WL&index=59
# from 35:00 onwards

class DataLayer():

    def __init__(self):
        self.filename = "BDEAllocationSheet.xlsx"
        self.is_loaded = False
        self.load_data()

    def load_data(self):
        if not self.is_loaded: # just bascially load it once
            self.book = openpyxl.load_workbook(self.filename)
            self.shFSP = self.book["FSPAlloc"]
            self.shIFA = self.book["IFAlloc"]
            self.shLookup = self.book["BDEMaster"]
            self.is_loaded = True

    def get_FSP_data(self):
        self.load_data()
        data = []
        for row in self.shFSP.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data
    
    def get_header(self):
        self.load_data()
        return [cell.value for cell in self.shFSP[1]]
    
    def get_BDE(self):
        table_address = self.shLookup.tables["BDEMasterLkp"].ref
        return  [cell[0].value for cell in self.shLookup[table_address]][1:]

    def _expand_table(self, sheet, table_name):
        table = sheet.tables[table_name]
        current_ref = table.ref
        start_cell, end_cell = current_ref.split(':')
        end_col = ''.join(filter(str.isalpha, end_cell))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        new_end_cell = f"{end_col}{end_row + 1}"
        table.ref = f"{start_cell}:{new_end_cell}"      

    def add_FSP_record(self, record, table_name = "Table1"):
        self.shFSP.append(record)
        self._expand_table(self.shFSP, table_name)  # Changed to "Table1"
        self.save()

    def add_IFA_record(self, record):
        self.shIFA.append(record)
        self.save()
    
    def delete_FSP_record(self, row_index):
        self.shFSP.delete_rows(row_index+1)
        self.save()

    def update_FSP_record_by_index(self, row_index, record):
        """
        Update a record by its row index (0-based, not counting header)
        row_index=0 is the first data row, row_index=1 is the second, etc.
        """
        excel_row = row_index + 2  # +1 for header, +1 for Excel 1-based indexing
        
        for col_index, value in enumerate(record):
            self.shFSP.cell(row=excel_row, column=col_index + 1).value = value
        
        self.save()
    
    def save(self):
        # so when the sheet is already open and you dont want to cause a write error
        try:
            self.book.save(self.filename)
        except PermissionError as e:
            raise RuntimeError("The Workbook is open by someone else.") from e


if __name__ == "__main__":
    data = DataLayer()
    pp.pprint(data.get_FSP_data())
    #pp.pprint(data.get_header())
    #pp.pprint(data.get_BDE())
    record = ["FSP666", "Lizdom", "Jan","20 feb 2020"]
    data.update_FSP_record_by_index(row_index=0, record=record)
    #data.add_FSP_record(record)
    #data.delete_FSP_record(3)
    pp.pprint(data.get_FSP_data())
    # 55:00
    