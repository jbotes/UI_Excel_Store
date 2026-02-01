import openpyxl

class DataLayer():

    def __init__(self):
        self.filename = "staff data.xlsx"
        self.is_loaded = False
        self.load_data()

    def load_data(self):
        if not self.is_loaded:
            self.book = openpyxl.load_workbook(self.filename)
            self.shStaff = self.book["Staff"]
            self.shLookup = self.book["Lookup"]
            self.is_loaded = True

    def save(self):
        try:
            self.book.save(self.filename)
        except PermissionError as e:
            raise RuntimeError("The workbook is open in Excel. Please close and try again") from e



    def delete_record(self,row_index):
        self.shStaff.delete_rows(row_index+1)
        self.save()
    
    def update_record(self,record):
        for row in self.shStaff.iter_rows(min_row=2):
            if row[0].value == record[0] and row[1].value == record[1]:
                for j,value in enumerate(record):
                    row[j].value = value 
                break
        self.save()

    # Data retrieval
    def get_data(self):
        self.load_data()
        data = []
        for row in self.shStaff.iter_rows(min_row=2,values_only=True):
            data.append(row)
        return data

    def get_header(self):
        self.load_data()
        return [cell.value for cell in self.shStaff[1]]

    def get_departments(self):
        table_department_address = self.shLookup.tables["tbDepartment"].ref
        return [cell[0].value for cell in self.shLookup[table_department_address]]

    def add_record(self, record):
        self.shStaff.append(record)
        self.save()
        
if __name__ == "__main__":
    
    import pprint as pp
    data = DataLayer()

    # Add record
    record = ["John","Smith","Part-Time","IT"]
    data.add_record(record)    
    print("\nData after add:")
    pp.pprint(data.get_data())
    
    # Delete record
    data.delete_record(1)
    print("\nData after delete")
    pp.pprint(data.get_data())

    # Update record
    record = ['Javier', 'Austin', 'Part-time', 'Marketing']
    data.update_record(record)
    print("\nData after update")
    pp.pprint(data.get_data())
    
    # Get header
    print("\nGet header")
    pp.pprint(data.get_header())

    # Get departments
    print("\nGet departments")
    pp.pprint(data.get_departments())




